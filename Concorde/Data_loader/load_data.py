import numpy as np
from openpyxl import load_workbook
from math import sqrt


def lecteur(fname):

    global wb, ws, n, dist, dem, dmax, a, b, alpha

    wb = load_workbook(fname)
    ws = wb["Feuil1"]
    n=0


    while  ws.cell(n+2,1).value is not None :

        n = n + 1

    dist=np.zeros([n,n]) #distancier
    dem = np.zeros(n) #demande
    tgg = np.zeros(n, dtype=int) #tournée géante
    a,b = np.zeros(n),np.zeros(n)
    dmax=0
    cox=np.zeros(n,dtype=float)
    coy=np.zeros(n,dtype=float)
    alpha=np.zeros(n, dtype=float)

    for i in range(n):
        alpha[i] = ws.cell(i+2,6).value


    for i in range(n):

        a[i] = ws.cell(i+2,4).value
        b[i] = ws.cell(i+2,5).value


    for i in range(n):
        cox[i]=ws.cell(2+i,2).value
        coy[i]=ws.cell(2+i,3).value
        dem[i]=ws.cell(2+i,7).value


    for i in range(n):
        for j in range(n):
            dist[i,j]=sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)
            dmax+=dist[i,j]

    return dist