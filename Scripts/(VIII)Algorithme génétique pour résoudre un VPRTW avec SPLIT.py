
#from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import time
from tqdm import tqdm
import math
import matplotlib.pyplot as plt
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

# Variables globales

m = 200 #population
prob_mut = 35 # pourcentage de mutation
perc_popu_mut = 0.4 # pourcentage de la population globale mutée
nbgen = 800 # nombre de générations


# Variable d'initialisation

percent_meilleure_insertion = round(m/100) #pourcentage de la population issue d'une meilleure insertion
percent_random = round(m/3) # pourcentage de la population choisie de manière random
percent_closest_neighbour = m - round(percent_meilleure_insertion) - round(percent_random) # pourcentage de la population choisie avec le voisin le plus proche


# Variable mutation 

pile_face = random.randint(0,1) # choisit la part de mutation ou on fait un swap randomisé( = 1) et une invertion complète de la population ( = 0)

# Variable de sélection

elite = 7 # nombre de meilleure solutions qu'on ne va pas muter (les 7 première solutions)
aug_popu = 1.12 # la population augmente de 1.12 % à chaque itérations



def lecteur():

    global wb, ws, n, dist, a, b, alpha, cox, coy, capa, dem, dmax

    wb = load_workbook(path_file)
    ws = wb["Feuil1"]
    n=0
    dmax = 0
    
    while  ws.cell(n+2,1).value is not None : 
        n=n+1
   
    dist=np.zeros([n,n], dtype=float)
    cox=np.zeros(n,dtype=float)
    coy=np.zeros(n,dtype=float)
    alpha=np.zeros(n, dtype=float)
    a = np.zeros(n, dtype = float)
    b = np.zeros(n, dtype = float)
    dem = np.zeros(n,dtype = float)

    capa = ws.cell(2,8).value
  

    #lire les distances dans la matrice dist
    
    for i in range(n):

        cox[i]=ws.cell(2+i,2).value
        coy[i]=ws.cell(2+i,3).value
        alpha[i] = ws.cell(2+i,6).value
        a[i] = ws.cell(2+i,4).value
        b[i] = ws.cell(2+i,5).value
        dem[i] = ws.cell(2+i,7).value


    for i in range(n):
        for j in range(n):
            dist[i,j]=math.sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)
            dmax += dist[i,j] 



def coutsN(route):

  cout = 0
  t = np.zeros(n, dtype = float)
  

  for i in range(len(route)-1):

    cout+=dist[route[i], route[i+1]]

    t[route[i+1]] = t[route[i]] + dist[route[i], route[i+1]]

    if t[route[i+1]] < a[route[i+1]]:
      t[route[i+1]] = a[route[i+1]]
 
    elif t[route[i+1]] > b[route[i+1]]:
      cout += (alpha[route[i+1]])*(t[route[i+1]]-b[route[i+1]])


  cout += dist[route[len(route)-2],0]
  return cout



def couts(route):

  cout = 0
  t = np.zeros(n, dtype = float)
  


  for i in range(len(route)):
    for j in range(len(route[i])-1): 
      
      cout+=dist[route[i][j], route[i][j+1]]

      t[route[i][j+1]] = t[route[i][j]] + dist[route[i][j], route[i][j+1]]


      if t[route[i][j+1]] < a[route[i][j+1]]:
        t[route[i][j+1]] = a[route[i][j+1]]
  
      elif t[route[i][j+1]] > b[route[i][j+1]]:
        cout += (alpha[route[i][j+1]])*(t[route[i][j+1]]-b[route[i][j+1]])

    cout += dist[route[i][len(route[i])-1],0]
  return cout


'''
def split(tourne_realisable):

    global temp_route,new_route

    V=np.full(n,dmax,dtype=float)
    P=np.zeros(n,dtype=int)
    V[0]=0

    for i in range(n - 1):
      
        
        D=0
        cout=0
        j = i+1

        while j<n and D <=capa:

            D += dem[tourne_realisable[j]]

            if j == i+1:
                new_route = [0,tourne_realisable[j],0]
                cout = coutsN(new_route)


            else:
              
                new_route.insert(j,tourne_realisable[j])
                temp_route = new_route.copy()
                del new_route[len(new_route)-2]
                cout += coutsN(temp_route)-coutsN(new_route) 

            if D<= capa:

                if V[i] +cout<V[j]:
                    V[j] = V[i] + cout
                    P[j]= i
                j=j+1

    solution=[]
    deb = P[n-1]
    fin = n-1

    while fin>0:

        route=[]
        route.append(0)

        for j in range(deb+1,fin+1):

            route.append(tourne_realisable[j])
        route.append(0)
        solution.append(route)

        fin=deb
        deb=P[fin]


    return solution'''



def split(tourne_realisable):

     V=np.full(n,dmax,dtype=float)
     P=np.zeros(n,dtype=int)
     V[0]=0

     for i in range(n - 1):
      
        
         D=0
         cout=0
         j = i+1

         while j<n and D <=capa:

             D += dem[tourne_realisable[j]]

             if j == i+1:
                 cout = dist[0,tourne_realisable[j]] + dist[tourne_realisable[j],0]

             else:

                 cout +=dist[tourne_realisable[j-1],tourne_realisable[j]]+dist[tourne_realisable[j],0]-dist[tourne_realisable[j-1],0]

             if D<= capa:

                 if V[i] +cout<V[j]:
                     V[j] = V[i] + cout
                     P[j]= i
                 j=j+1

     solution=[]
     deb = P[n-1]
     fin = n-1

     while fin>0:

         route=[]
         route.append(0)

         for j in range(deb+1,fin+1):

             route.append(tourne_realisable[j])
         route.append(0)
         solution.append(route)

         fin=deb
         deb=P[fin]

     return solution



def rand_two_closest(r,visited):    
  
    v1=0
    dref=10000
    for i in range(1,n):
        if(dist[r,i]<dref and visited[i]==False):
            dref=dist[r,i]
            v1=i
    v2=-1
    dref=10000
    for i in range(1,n):
        if(dist[r,i]<dref and visited[i]==False and i!=v1):
            dref=dist[r,i]
            v2=i
    a=random.randint(0,1)
    if v2==-1 or a==1:
        return v1
    else : 
        return v2



def closest_neighbour():
    
    visited = np.full(n,False,dtype=bool)
    route=np.zeros(n,dtype=int)
    #construire une route avec l'aide de rand_two_closest
    
    route[0]=0
    visited[0]=True
    
    for i in range(n-1):
        
        r = rand_two_closest(route[i],visited)
        route[i+1]=r
        visited[r]=True

    route = list(route)
    route.append(0)

    return route



def meilleure_insertion():

  insertion1 = random.randint(1,n-1)

  r=[0,insertion1,0]


  while len(r) != (n+1):

    min_liste = np.zeros((len(r)-1,2))

    for i in range(len(r)-1):

      liste = [dist[r[i],p] for p in range(n) ]
      minimum = np.min([dist[r[i],p] for p in range(n) if p not in r])


      for j in range(len(liste)):
        if liste[j] == minimum:
          if j not in r:
            index = j

      min_liste[i]=[minimum,index]


    for k in range(len(r)-1):
      if min_liste[k,0] == np.min(min_liste[:,0]):
        index2=min_liste[k,1]
        position = k

    r.insert(position+1,int(index2))

  return r



def population_random():

  route=[0]

  while len(route)!= n:

    sommet = random.randint(0,n-1)
    if sommet not in route:
      route.append(sommet)
    

  route.append(0)

  return route



def initialize():

  global population

  population = {}

  #initialiser la population random

  for indice in range(round(percent_random)):
    route = population_random()
    routeVRP = split(route)
    population[couts(routeVRP)] = route, routeVRP


  #initialiser la poplation avec meilleure insertion

  while len(population) != (round(percent_meilleure_insertion) +round(percent_random)):
    route = meilleure_insertion()
    routeVRP = split(route)
    population[couts(routeVRP)] = route, routeVRP

  # initialiser la population avec closest_neighbour 

  for indice in range(round(percent_closest_neighbour)):
    route = closest_neighbour()
    routeVRP = split(route)
    population[couts(routeVRP)] = route, routeVRP

  return population



def mutation(route):


  if pile_face == 0:

    mut1 = random.randint(1,len(route)-1)
    mut2 = (len(route))-mut1 

    while mut1 == mut2:
      mut1 = random.randint(1,len(route)-1)
      mut2 = (len(route))-mut1 

    if mut1 > mut2 : 
      temp1 = route[mut1:len(route)-1]
      temp2 = route[1:mut2]

    if mut2 > mut1:
      temp2 = route[mut1:len(route)-1]
      temp1 = route[1:mut2]



    route[mut1:len(route)-1] = temp2
    route[1:mut2] = temp1
    return route
  


  if pile_face == 1:

      nb_change = random.randint(1,n/5)

      for i in range(nb_change):

        position_change1 = random.randint(1, len(route)-2)
        position_change2 = random.randint(1,len(route)-2)

        if route[position_change1] != route[position_change2]:
          a=route[position_change1]
          b=route[position_change2]

          route[position_change2]=a
          route[position_change1]=b

      return route


def cross_over(p1,p2):

# On initialise les enfants

  enfant1 = list(np.zeros(len(p1),dtype=int))
  enfant2 = list(np.zeros(len(p1), dtype = int))

  rliste = [i for i in range(n)]

  part_p = random.randint(0,len(p1))


  enfant1[:part_p] = p1[:part_p]
  enfant1[part_p:] = p2[part_p:]

  enfant2[:part_p] = p2[:part_p]
  enfant2[part_p:] = p1[part_p:]


  for i in range(1,len(enfant1)-1):

    if list(enfant1).count(enfant1[i]) > 1: 
        for r in rliste :
          if r not in enfant1 :
            enfant1[i] = r


  for i in range(1,len(enfant2)-1):

    if list(enfant2).count(enfant2[i]) > 1 : 

        for r in rliste:
          if r not in enfant2 :
            enfant2[i] = r

  return list(enfant1) , list(enfant2)



def selection(population):

  population = dict(population)
  population =  sorted(population.items(), key=lambda t: t[0])
  

  return population[0:round(aug_popu*m)+m]



def main():    
  global population, bcosts

  lecteur()
  initialize()
  bcosts=np.zeros(nbgen,dtype=float)
  avgcosts=np.zeros(nbgen,dtype=float)
  begin = time.time()
  
  
  population = [(k, v) for k, v in population.items()]   

  for gen in tqdm(range(nbgen), desc="Generation", unit="gen", unit_scale=True,
                bar_format="{desc}: {percentage:3.0f}% |{bar}| {n_fmt}/{total_fmt} {elapsed} <remaining time: {remaining}, rate: {rate_fmt}>",
                dynamic_ncols=True, smoothing=0.5):


    for i in range(round(m)):



      parent1 = list(population[random.randint(0,m-1)][1][0])
      parent2 = list(population[random.randint(0,m-1)][1][0])


      if  parent1 == parent2: 
        parent2 = list(population[random.randint(0,m-1)][1][0]) 
      
      enfant = cross_over(parent1,parent2) 


      population.append((couts(split(enfant[0])),(enfant[0], split(enfant[0]))))
      population.append((couts(split(enfant[1])),(enfant[1], split(enfant[1]))))


    if (random.randint(1, round(len(population)/prob_mut)) == 1):

      for i in range(round(perc_popu_mut*len(population))):


        random_int = random.randint(elite,len(population)-1)
        temp = mutation(population[random_int][1][0])

        population.append((couts(split(temp)), (temp, split(temp))))



    population = selection(population)


    bcosts[gen] = population[0][0]
    avgcosts[gen] = sum(population[i][0] for i in range(len(population)))/len(population)

  end=time.time()
  elapsed=end-begin

  pd.DataFrame({'best': bcosts,'avg': avgcosts}).plot()
  pd.DataFrame(avgcosts-bcosts).plot()
  pd.DataFrame(bcosts).plot()
  print(bcosts[nbgen-1], population[0][1][1], len(population))
  print('Temps d\'exécution : {:.3f}s - temps moy \'exécution d\'une itération : {:.3f} '.format(elapsed,elapsed/nbgen))

  plt.show()

if __name__ == '__main__':
    main()




import matplotlib.pyplot as plt
import networkx as nx
global pos
pos={}

for i in range(n):
    pos[i]=np.array([cox[i],coy[i]])


def graphh(route):
    g=nx.Graph()
    g.add_nodes_from({i for i in range(n)})
    for i in range(len(route)):
      for j in range(len(route[i])-1):
        g.add_edges_from({(route[i][j],route[i][j+1])})
    plt.figure(figsize=(10,10))
    nx.draw_networkx(g,pos)
    plt.title('meilleur solution')
    plt.show()


graphh(population[0][1][1])