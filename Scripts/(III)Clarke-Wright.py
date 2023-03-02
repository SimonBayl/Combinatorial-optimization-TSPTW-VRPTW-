from ast import While
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import itertools as iter
import time
import math
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

wb = load_workbook(path_file)
ws = wb["Feuil1"]
n=0

while  ws.cell(n+2,1).value is not None : 
    n=n+1

dist=np.zeros([n,n], dtype=float)
cox=np.zeros(n,dtype=float)
coy=np.zeros(n,dtype=float)
alpha=np.zeros(n, dtype=float)
a = np.zeros(n, dtype = float)
b = np.zeros(n, dtype = float)

#lire les distances dans la matrice dist

for i in range(n):

    cox[i]=ws.cell(2+i,2).value
    coy[i]=ws.cell(2+i,3).value
    alpha[i] = ws.cell(2+i,6).value
    a[i] = ws.cell(2+i,4).value
    b[i] = ws.cell(2+i,5).value
    

for i in range(n):
    for j in range(n):
        dist[i,j]=math.sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)



def CW():

  route =[]
  visited = []
  c_route = []
  solution = []
  max = 0

  while len(route) != n-1:

    a = random.randint(1,n-1)

    if a not in visited:
      route.append([0,a,0])
      visited.append(a)

  for i in range(len(route)):
    for j in range(len(route)):
      if i!=j and dist[0,route[i][1]]+dist[route[j][1],0] - dist[route[i][1],route[j][1]] not in [c_route[i][0] for i in range(len(c_route))]:
        c_route.append([dist[0,route[i][1]]+dist[route[j][1],0] - dist[route[i][1],route[j][1]], [0,route[i][1],route[j][1],0]])

  c_route.sort(reverse = True)
  i=0

  while i != len(route)-1:
    if i < (n-1)-3 :
      for j in range(1,len(route[i])-1):
        if route[i][j] in c_route[0][1]:
          del route[i]
        
    i+=1

  route.append(c_route[0][1])
  route_init = c_route[0][1]

  amelioration = True 
  
  while amelioration == True:
    amelioration = False 
    
    c_route = []
    for j in range(len(route)-1):
      if route[j][-2] not in route_init:
        c_route.append([float(dist[0,route_init[1]] + dist[route[j][-2],0] - dist[route[j][-2],route_init[1]]), [route[j][:len(route[j])-1]+route_init[1:]]])
        c_route.append([float(dist[route_init[-2],0] + dist[0,route[j][1]] - dist[route[j][1],route_init[-2]]), [route_init[:len(route_init)-1]+route[j][1:]]])

    c_route.sort(reverse = True)

    for i in range(len(route)-1):

      if i >= len(route):
        break

      for j in range(1,(len(route[i])-1)):

        if route[i][j] in c_route[0][1][0]:
          del route[i]

          if i == len(route):
            break

    if len(route[-1]) > 3:
      del route[-1]


    if c_route != [] :
      route.append(c_route[0][1][0])
      route_init = c_route[0][1][0]
      amelioration = True 

    if len(c_route[0][1][0]) == n+1:
      route =[]
      route.append(c_route[0][1][0])
      amelioration = False

    i+=1

  route = route[0]
  return route



def couts(route):
  cout = 0

  for i in range(len(route)-2):
    cout+= dist[route[i],route[i+1]]
  cout+= dist[route[n-1],0]
  return cout


import matplotlib.pyplot as plt
import networkx as nx
global pos
pos={}

for i in range(n):
    pos[i]=np.array([cox[i],coy[i]])

print(CW())
# graphique(CW())

