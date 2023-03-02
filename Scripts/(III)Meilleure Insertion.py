
from ast import While
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import itertools as iter
import time
import math
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

wb = load_workbook(path_file)
ws = wb["Feuil1"]
n=0


while  ws.cell(n+2,1).value is not None : 
    n=n+1

dist=np.zeros([n,n], dtype=float)
cox=np.zeros(n,dtype=float)
coy=np.zeros(n,dtype=float)
alpha=np.zeros(n, dtype=float)
a = np.zeros(n, dtype = float)
b = np.zeros(n, dtype = float)

#lire les distances dans la matrice dist

for i in range(n):

    cox[i]=ws.cell(2+i,2).value
    coy[i]=ws.cell(2+i,3).value
    alpha[i] = ws.cell(2+i,6).value
    a[i] = ws.cell(2+i,4).value
    b[i] = ws.cell(2+i,5).value
    

for i in range(n):
    for j in range(n):
        dist[i,j]=math.sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)



def meilleure_insertion():

  insertion1 = random.randint(1,n-1)

  r=[0,insertion1,0]

  while len(r) != (n+1):

    min_liste = np.zeros((len(r)-1,2))

    liste={}
    for i in range(len(r)-1):

      for p in range(n):
        if p not in r:
          var_rand = random.randint(0,1)
          if var_rand == 0 or i <=1:
            r.insert(i+1,p)
            liste[couts(r)] = [p, i+1]
          if var_rand == 1 and i > 1 :
            r.insert(i-1,p)
            liste[couts(r)] = [p, i-1]

          r.remove(p)

    liste =  sorted(liste.items(), key=lambda t: t[0])

    if liste !=[]:
      r.insert(liste[0][1][1], liste[0][1][0])
  
  return r

def couts(route):

  cout = 0
  t = np.zeros(n, dtype = float)
  

  for i in range(len(route)-1):

    cout+=dist[route[i], route[i+1]]

    t[route[i+1]] = t[route[i]] + dist[route[i], route[i+1]]

    if t[route[i+1]] < a[route[i+1]]:
      t[route[i+1]] = a[route[i+1]]
 
    elif t[route[i+1]] > b[route[i+1]]:
      cout += (alpha[route[i+1]])*(t[route[i+1]]-b[route[i+1]])


  cout += dist[route[len(route)-2],0]
  return cout



import matplotlib.pyplot as plt
import networkx as nx
global pos
pos={}

for i in range(n):
    pos[i]=np.array([cox[i],coy[i]])



def graphique(route):
    g=nx.Graph()
    g.add_nodes_from({i for i in range(n)})
    for i in range(len(route)-1):
        g.add_edges_from({(route[i],route[i+1])})
    # g.add_edges_from({(route[-1],0)})
    plt.figure(figsize=(10,10))
    nx.draw_networkx(g,pos)
    plt.title('meilleure solution')
    plt.show()

graphique(meilleure_insertion())

print(couts(meilleure_insertion()))
print(meilleure_insertion())