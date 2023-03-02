#from openpyxl import Workbook
from openpyxl import load_workbook
import math
import pandas as pd
import numpy as np
import random
import time
from tqdm import tqdm
import matplotlib.pyplot as plt
import statistics
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

# Variables globales

m = 110 # 700  400 population
prob_mut = 40 # 45 30 pourcentage de mutation
perc_popu_mut = 0.45 # 0.45 0.3 pourcentage de la population globale mutée
nbgen = 700 # 700  1000 nombre de générations


# Variable d'initialisation

percent_meilleure_insertion =m/10 # m/100  pourcentage de la population issue d'une meilleure insertion
percent_random =  m/3 # m/3 pourcentage de la population choisie de manière random
percent_clarke_wright = m/4 # m/10 m/5 pourcentage de la population choisie avec un Clarke and Wright
percent_closest_neighbour =  m - round(percent_clarke_wright) - round(percent_meilleure_insertion) - round(percent_random) # pourcentage de la population choisie avec le voisin le plus proche


# Variable mutation 

pile_face = 1 # 1 random.randint(0,1) # choisit la part de mutation ou on fait un swap randomisé( = 1) et une invertion complète de la population ( = 0)

# Variable de sélection

elite = 1 # 1 nombre de meilleure solutions qu'on ne va pas muter (les 7 première solutions)
aug_popu = 1 # 1 la population augmente de 1.12 % à chaque itérations



def lecteur():

    global wb, ws, n, dist, a, b, alpha, cox, coy

    wb = load_workbook(path_file)
    ws = wb["Feuil1"]
    n=0
    
    while  ws.cell(n+2,1).value is not None : 
        n=n+1
   
    dist=np.zeros([n,n], dtype=float)
    cox=np.zeros(n,dtype=float)
    coy=np.zeros(n,dtype=float)
    alpha=np.zeros(n, dtype=float)
    a = np.zeros(n, dtype = float)
    b = np.zeros(n, dtype = float)
    
    #lire les distances dans la matrice dist
    
    for i in range(n):

        cox[i]=ws.cell(2+i,2).value
        coy[i]=ws.cell(2+i,3).value
        alpha[i] = ws.cell(2+i,6).value
        a[i] = ws.cell(2+i,4).value
        b[i] = ws.cell(2+i,5).value

    for i in range(n):
        for j in range(n):
            dist[i,j]=math.sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)



def couts(route):

  cout = 0
  t = np.zeros(n, dtype = float)
  

  for i in range(len(route)-1):

    cout+=dist[route[i], route[i+1]]

    t[route[i+1]] = t[route[i]] + dist[route[i], route[i+1]]

    if t[route[i+1]] < a[route[i+1]]:
      t[route[i+1]] = a[route[i+1]]
 
    elif t[route[i+1]] > b[route[i+1]]:
      cout += (alpha[route[i+1]])*(t[route[i+1]]-b[route[i+1]])


  cout += dist[route[len(route)-2],0]
  return cout



def Clarke_Wright():
  
  route =[]
  visited = []
  c_route = []
  solution = []
  max = 0

  while len(route) != n-1:

    a = random.randint(1,n-1)

    if a not in visited:
      route.append([0,a,0])
      visited.append(a)

  for i in range(len(route)):
    for j in range(len(route)):
      if i!=j and dist[0,route[i][1]]+dist[route[j][1],0] - dist[route[i][1],route[j][1]] not in [c_route[i][0] for i in range(len(c_route))]:
        c_route.append([dist[0,route[i][1]]+dist[route[j][1],0] - dist[route[i][1],route[j][1]], [0,route[i][1],route[j][1],0]])

  # c_route.sort(reverse = True)   #Sert a passer en mode random si commenté sinon mode non random
  i=0

  while i != len(route)-1:
    if i < (n-1)-3 :
      for j in range(1,len(route[i])-1):
        if route[i][j] in c_route[0][1]:
          del route[i]
        
    i+=1

  route.append(c_route[0][1])
  route_init = c_route[0][1]

  amelioration = True 
  while amelioration == True:
    amelioration = False 
    
    c_route = []
    for j in range(len(route)-1):
      if route[j][-2] not in route_init:
        c_route.append([float(dist[0,route_init[1]] + dist[route[j][-2],0] - dist[route[j][-2],route_init[1]]), [route[j][:len(route[j])-1]+route_init[1:]]])
        c_route.append([float(dist[route_init[-2],0] + dist[0,route[j][1]] - dist[route[j][1],route_init[-2]]), [route_init[:len(route_init)-1]+route[j][1:]]])

    c_route.sort(reverse = True)

    for i in range(len(route)-1):

      if i >= len(route):
        break

      for j in range(1,(len(route[i])-1)):

        if route[i][j] in c_route[0][1][0]:
          del route[i]

          if i == len(route):
            break

    if len(route[-1]) > 3:
      del route[-1]


    if c_route != [] :
      route.append(c_route[0][1][0])
      route_init = c_route[0][1][0]
      amelioration = True 

    if len(c_route[0][1][0]) == n+1:
      route =[]
      route.append(c_route[0][1][0])
      amelioration = False

  route = route[0]
  return route



def rand_two_closest(r,visited):    
  
    v1=0
    dref=10000

    for i in range(1,n):
        if(dist[r,i]<dref and visited[i]==False):
            dref=dist[r,i]
            v1=i
    v2=-1
    dref=10000
    for i in range(1,n):
        if(dist[r,i]<dref and visited[i]==False and i!=v1):
            dref=dist[r,i]
            v2=i
    a=random.randint(0,1)
    if v2==-1 or a==1:
        return v1
    else : 
        return v2



def closest_neighbour():
    
    visited = np.full(n,False,dtype=bool)
    route=np.zeros(n,dtype=int)
    #construire une route avec l'aide de rand_two_closest
    
    route[0]=0
    visited[0]=True
    
    for i in range(n-1):
        
        r = rand_two_closest(route[i],visited)
        route[i+1]=r
        visited[r]=True

    route = list(route)
    route.append(0)
  
    return route


def meilleure_insertion():


  insertion1 = random.randint(1,n-1)

  r=[0,insertion1,0]


  while len(r) != (n+1):

    min_liste = np.zeros((len(r)-1,2))

    liste={}
    for i in range(len(r)-1):

      for p in range(n):
        if p not in r:
          var_rand = random.randint(0,1)
          if var_rand == 0 or i <=1:
            r.insert(i+1,p)
            liste[couts(r)] = [p, i+1]
          if var_rand == 1 and i > 1 :
            r.insert(i-1,p)
            liste[couts(r)] = [p, i-1]

          r.remove(p)

    liste =  sorted(liste.items(), key=lambda t: t[0])


    if liste !=[]:
      r.insert(liste[0][1][1], liste[0][1][0])
    

  return r



def population_random():

  route=[0]

  while len(route)!= n:

    sommet = random.randint(0,n-1)
    if sommet not in route:
      route.append(sommet)
    

  route.append(0)

  return route



def initialize():

  global population

  population = {}

  #initialiser la population random

  for indice in range(round(percent_random)):
    route = population_random()
    population[couts(route)] = route

  #initialiser la poplation avec meilleure insertion

  while len(population) != (round(percent_meilleure_insertion) +round(percent_random)):
    route = meilleure_insertion()
    population[couts(route)] = route

  # initialiser la population avec closest_neighbour 

  for indice in range(round(percent_closest_neighbour)):
    route = closest_neighbour()
    population[couts(route)] = route

  # initialiser la population avec un Clarke and Wright

  for indice in range(round(percent_clarke_wright)):
    route = Clarke_Wright()
    population[couts(route)] = route

  return population



def mutation(route):

  # pil_face = random.randint(0,1)

  if pile_face == 0:

    mut1 = random.randint(1,len(route)-1)
    mut2 = (len(route))-mut1 

    while mut1 == mut2:
      mut1 = random.randint(1,len(route)-1)
      mut2 = (len(route))-mut1 

    if mut1 > mut2 : 
      temp1 = route[mut1:len(route)-1]
      temp2 = route[1:mut2]

    if mut2 > mut1:
      temp2 = route[mut1:len(route)-1]
      temp1 = route[1:mut2]

    route[mut1:len(route)-1] = temp2
    route[1:mut2] = temp1
    return route

  if pile_face == 1:

      nb_change = random.randint(1,round(n/5))

      for i in range(nb_change):

        position_change1 = random.randint(1, len(route)-2)
        position_change2 = random.randint(1,len(route)-2)

        if route[position_change1] != route[position_change2]:
          a=route[position_change1]
          b=route[position_change2]

          route[position_change2]=a
          route[position_change1]=b

      return route



def cross_over(p1,p2):

# On initialise les enfants

  enfant1 = list(np.zeros(len(p1),dtype=int))
  enfant2 = list(np.zeros(len(p1), dtype = int))

# On définit les coefficient pour que la liste random soit plus ou moins influencée par un parent
  
  rliste = [i for i in range(n)]

  part_p = random.randint(0,len(p1))

  enfant1[:part_p] = p1[:part_p]
  enfant1[part_p:] = p2[part_p:]

  enfant2[:part_p] = p2[:part_p]
  enfant2[part_p:] = p1[part_p:]


  for i in range(1,len(enfant1)-1):

    if list(enfant1).count(enfant1[i]) > 1: 
        for r in rliste :
          if r not in enfant1 :
            enfant1[i] = r


  for i in range(1,len(enfant2)-1):

    if list(enfant2).count(enfant2[i]) > 1 : 

        for r in rliste:
          if r not in enfant2 :
            enfant2[i] = r

  return list(enfant1) , list(enfant2)



def selection(population):

  population = dict(population)
  population =  sorted(population.items(), key=lambda t: t[0])

  return population[0:m]



def two_opt_TW(route):

  amelioration = True 
  while amelioration == True:
    amelioration = False

    for i in range(1,len(route)-2):
      for j in range(1,len(route)-2):
        if j != i-1 and j != i and j != i+1:
          
          route_TW = np.zeros(len(route), dtype = int)
          for k in range(len(route)):
            route_TW[k] = route[k]

          route_TW[i+1] = route[j]
          route_TW[j] = route[i+1]
          route_TW[i+2:j] = list(reversed(route[i+2:j]))
          

          if couts(route) > couts(route_TW):
            route = route_TW

            amelioration = True

  return list(route)



def main():    
  global population, bcosts, solution

  lecteur()
  initialize()

  bcosts =[]

  avgcosts =[]
  begin = time.time()
  
  
  population = [(k, v) for k, v in population.items()]   
  
  for gen in tqdm(range(nbgen), desc="Generation", unit="gen", unit_scale=True,
                bar_format="{desc}: {percentage:3.0f}% |{bar}| {n_fmt}/{total_fmt} {elapsed} <remaining time: {remaining}, rate: {rate_fmt}>",
                dynamic_ncols=True, smoothing=0.5):


    for i in range(round(m)):
      parent1 = list(population[random.randint(0,m-1)][1])
      parent2 = list(population[random.randint(0,m-1)][1])

      if  parent1 == parent2: 
        parent2 = list(population[random.randint(0,m-1)][1]) 
      
      enfant = cross_over(parent1,parent2) 
      population.append((couts(enfant[0]),enfant[0]))
      population.append((couts(enfant[1]),enfant[1]))

    if (random.randint(1, round(len(population)/prob_mut)) == 1):

      for i in range(round(perc_popu_mut*len(population))):

        random_int = random.randint(elite,len(population)-1)
        temp = mutation(population[random_int][1])
        population.append((couts(temp), temp))

    population = selection(population)
    bcosts.append(population[0][0])
    avgcosts.append(statistics.median([population[i][0] for i in range(len(population))] ))

  end=time.time()
  elapsed=end-begin

  # Recherche locale avec 2-opt

  solution = two_opt_TW(population[0][1])
  cout_soluce = couts(solution)

  if cout_soluce < bcosts[-1]:
    bcosts.append(couts(solution))

    avgcosts.append(statistics.median([population[i][0] for i in range(len(population))]))
  else:
    solution = population[0][1]

  print(bcosts[-1], solution)
  pd.DataFrame({'best': bcosts[10:],'avg': avgcosts[10:]}).plot()
  plt.show()
  plt.plot(bcosts)

  print('Temps d\'exécution : {:.3f}s - temps moy \'exécution d\'une itération : {:.3f} '.format(elapsed,elapsed/nbgen))
  plt.show()

if __name__ == '__main__':
    main()

import matplotlib.pyplot as plt
import networkx as nx
global pos
pos={}

for i in range(n):
    pos[i]=np.array([cox[i],coy[i]])



def graphique(route):
    g=nx.Graph()
    g.add_nodes_from({i for i in range(n)})
    for i in range(len(route)-1):
        g.add_edges_from({(route[i],route[i+1])})

    plt.figure(figsize=(10,10))
    nx.draw_networkx(g,pos)
    plt.title('meilleure solution')
    plt.show()

graphique(solution)