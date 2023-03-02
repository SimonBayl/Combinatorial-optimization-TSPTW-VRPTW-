from ast import While
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import itertools as iter
import time
import math
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

wb = load_workbook(path_file)
ws = wb["Feuil1"]
n=0

while  ws.cell(n+2,1).value is not None : 
    n=n+1

dist=np.zeros([n,n], dtype=float)
cox=np.zeros(n,dtype=float)
coy=np.zeros(n,dtype=float)
alpha=np.zeros(n, dtype=float)
a = np.zeros(n, dtype = float)
b = np.zeros(n, dtype = float)


#lire les distances dans la matrice dist

for i in range(n):

    cox[i]=ws.cell(2+i,2).value
    coy[i]=ws.cell(2+i,3).value
    alpha[i] = ws.cell(2+i,6).value
    a[i] = ws.cell(2+i,4).value
    b[i] = ws.cell(2+i,5).value
    

for i in range(n):
    for j in range(n):
        dist[i,j]=math.sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)



def two_opt(route):

  amelioration = True 
  while amelioration == True:
    amelioration = False

    for i in range(1,len(route)-2):
      for j in range(1,len(route)-2):
        if j != i-1 and j != i and j != i+1:

          if dist[route[i], route[i+1]] + dist[route[j], route[j+1]] > dist[route[i], route[j]] + dist[route[i+1], route[j+1]]:
            temp1 = route[i+1]
            temp2 = route[j]
            route[i+1]=temp2
            route[j]=temp1
            route[i+2:j] = list(reversed(route[i+2:j]))

            amelioration = True
  print("2opt", route)
  return route



def two_opt_TW(route):

  amelioration = True 
  while amelioration == True:
    amelioration = False

    for i in range(1,len(route)-2):
      for j in range(1,len(route)-2):
        if j != i-1 and j != i and j != i+1:
          
          route_TW = np.zeros(len(route), dtype = int)
          for k in range(len(route)):
            route_TW[k] = route[k]

          route_TW[i+1] = route[j]
          route_TW[j] = route[i+1]
          route_TW[i+2:j] = list(reversed(route[i+2:j]))
          

          if coutsTW(route) > coutsTW(route_TW):
            route = route_TW

            amelioration = True
  print("2optTW", route)
  return list(route)




route = [0, 12, 11, 8, 32, 39, 35, 31, 27, 29, 4, 13, 18, 37, 25, 17, 5, 15, 23, 36, 14, 7, 10, 34, 6, 22, 28, 38, 21, 20, 30, 16, 24, 33, 19, 2, 26, 9, 3, 1, 0]



def couts(route):

  cout = 0

  for i in range(len(route)-2):
    cout+= dist[route[i],route[i+1]]
  cout+= dist[route[n-1],0]
  return cout, len(route)



def coutsTW(route):

  cout = 0
  t = np.zeros(n, dtype = float)
  

  for i in range(len(route)-1):

    cout+=dist[route[i], route[i+1]]

    t[route[i+1]] = t[route[i]] + dist[route[i], route[i+1]]

    if t[route[i+1]] < a[route[i+1]]:
      t[route[i+1]] = a[route[i+1]]
 
    elif t[route[i+1]] > b[route[i+1]]:
      cout += (alpha[route[i+1]])*(t[route[i+1]]-b[route[i+1]])


  cout += dist[route[len(route)-1],0]
  return cout



import matplotlib.pyplot as plt
import networkx as nx
global pos
pos={}

for i in range(n):
    pos[i]=np.array([cox[i],coy[i]])



def graphique(route):
    g=nx.Graph()
    g.add_nodes_from({i for i in range(n)})
    for i in range(len(route)-1):
        g.add_edges_from({(route[i],route[i+1])})
    plt.figure(figsize=(10,10))
    nx.draw_networkx(g,pos)
    plt.title('meilleure solution')
    plt.show()

graphique(route)
graphique(two_opt(route))