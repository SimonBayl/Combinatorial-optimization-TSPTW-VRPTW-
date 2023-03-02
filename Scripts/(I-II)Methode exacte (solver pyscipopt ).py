
"""
Created on Mon Sep 20 14:29:13 2021

@author: bayle/chesneau
"""

from openpyxl import load_workbook
from pyscipopt import Model, quicksum,Expr
import numpy as np
from math import sqrt
import os 

path_file = os.getcwd() + '\Sujet-Données\Instance.xlsx'

def lecteur(fname):

    global wb, ws, n, dist, dem, dmax, a, b, alpha

    wb = load_workbook(fname)
    ws = wb["Feuil1"]
    n=0
    

    while  ws.cell(n+2,1).value is not None :

        n = n + 1

    dist=np.zeros([n,n]) #distancier
    dem = np.zeros(n) #demande
    tgg = np.zeros(n, dtype=int) #tournée géante
    a,b = np.zeros(n),np.zeros(n)
    dmax=0
    cox=np.zeros(n,dtype=float)
    coy=np.zeros(n,dtype=float)
    alpha=np.zeros(n, dtype=float)

    for i in range(n):
        alpha[i] = ws.cell(i+2,6).value


    for i in range(n):

        a[i] = ws.cell(i+2,4).value
        b[i] = ws.cell(i+2,5).value


    for i in range(n):
        cox[i]=ws.cell(2+i,2).value
        coy[i]=ws.cell(2+i,3).value
        dem[i]=ws.cell(2+i,7).value


    for i in range(n):
        for j in range(n):
            dist[i,j]=sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)
            dmax+=dist[i,j]


def LinMod():
    model = Model("Methode_exacte")  
    x={}
    t={}
    for i in range(n):
        t[i]=model.addVar(lb=0,ub=dmax**2,name="t[%s]"%i)
        for j in range(n):
            x[i,j]=model.addVar(vtype="B",name="x[%s,%s]"%(i,j))
    D={}
    for i in range(n):
      D[i]=model.addVar(lb=0,ub=dmax**2,name="t[%s]"%i)

    model.setObjective(quicksum(x[i,j]*dist[i,j] for i in range(n) for j in range(n)) + quicksum(alpha[i]*D[i] for i in range(n)) ,"minimize")

    mycst = Expr()

    for j in range(n):
        mycst += x[0,j]
    model.addCons(mycst==1)

    for j in range(n):
        model.addCons(quicksum(x[i,j] for i in range(n) if i!=j) == 1,"Visite %s"%j)

    for j in range(n):
      model.addCons(quicksum(x[i,j] for i in range(n)if i!=j) - quicksum(x[j,i]for i in range(n)if i!=j) ==0,"Cont %s"%j)

    for i in range(n):
      model.addCons(D[i]>=0, "date_arrivée")


    for i in range(n):
      model.addCons(D[i]>=t[i]-b[i], "date_arrivée")

    #MTZ
    for j in range(1,n):
      for i in range(n):
          if(i!=j):model.addCons(t[j]>=t[i]+dist[i,j]-4*dmax*(1-x[i,j]))
          
    for j in range(n):
      model.addCons(t[j]>=a[j])



    model.addCons(t[0]==0)

    # model.hideOutput()
    # Ligne à changer en fonction du temps d'éxécution souhaité
    # model.setParam("limits/time",7200)

    model.optimize()

    sol = model.getBestSol()

    print("Fonction obj %s "%model.getSolObjVal(sol))
    print("Temps de resolution", model.getSolvingTime() )
    print("Variables :")
    for i in range(n):
        for j in range(n):
            if model.getSolVal(sol,x[i,j])>0.99:
                print(i,"->",j," :: ",model.getSolVal(sol,t[i]))


    print("Tournee commencant par depot ")
    ori=0
    dest=-1
    while dest!=0:
        dest=0
        goon=True
        while dest<n and goon :
            if(model.getSolVal(sol,x[ori,dest])>0.5) :
                print(ori,"->",dest)
                ori=dest
                goon=False
            else:dest+=1



def main():
    lecteur(path_file)
    LinMod()


if __name__=="__main__": # la fonction main est lancée avec cette expression particulière
    main()